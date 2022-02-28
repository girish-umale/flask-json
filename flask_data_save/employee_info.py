import json
import openpyxl


class Employee:
    def __init__(self, empID, empName, empMail, empGender, fileType=None):
        self.empID = empID
        self.empName = empName
        self.empMail = empMail
        self.empGender = empGender
        self.fileType = fileType

    def __str__(self):
        return f'''(Eid : {self.empID}, Enm : {self.empName}, Email : {self.empMail}, Egender : {self.empGender})'''

    def __repr__(self):
        return str(self)


FILE_PATH = 'D:\\PythonProjects\\Flask_Practice\\flask_data_save\\data\\'
JSON_FILE_PATH = FILE_PATH + "emp.json"
TEXT_FILE_PATH = FILE_PATH + "emp.txt"
EXCEL_FILE_PATH = FILE_PATH + "emp.xlsx"
CSV_FILE_PATH = FILE_PATH + "emp.csv"
emplist = []


def write_into_txt(emp):
    with open(TEXT_FILE_PATH, 'a') as file:
        empstr = str(emp.empID) + "\t\t" + emp.empName + "\t\t" + emp.empMail + "\t\t" + emp.empGender + "\n"
        file.writelines(empstr)


def write_into_json(emp):
    with open(JSON_FILE_PATH, 'a') as file:
        json.dump(emp.__dict__, file)
        file.writelines('\n')


def write_into_csv(emp):
    with open(CSV_FILE_PATH, 'a') as file:
        empstr = str(emp.empID) + "," + emp.empName + "," + emp.empMail + "," + emp.empGender + ","
        file.writelines(empstr)


def write_into_excel(emp):
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook['emp_data']
        lastrow = sheet.max_row + 1
    except:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet('emp_data')
        lastrow = 1

    sheet['A' + str(lastrow)] = emp.empID
    sheet['B' + str(lastrow)] = emp.empName
    sheet['C' + str(lastrow)] = emp.empMail
    sheet['D' + str(lastrow)] = emp.empGender
    workbook.save(EXCEL_FILE_PATH)


def write_into_sqlite3(emp):
    pass


def read_from_json_file():
    emplist = []
    with open(JSON_FILE_PATH, 'r') as file:
        alllines = file.readlines()
        for line in alllines:
            emp = Employee(**json.loads(line))
            print(emp)
            emplist.append(emp)
    return emplist


FILE_TYPES_FUN_REF = {
    "T": write_into_txt,
    "J": write_into_json,
    "C": write_into_csv,
    "X": write_into_excel,
    "S": write_into_sqlite3
}
